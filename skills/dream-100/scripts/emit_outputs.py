#!/usr/bin/env python3
"""Writes ALL outputs of a Dream 100 run in one call:
  1. the formatted xlsx  ->  <base>/[C] Dream 100 Leads - <date>.xlsx
  2. client DB append    ->  <base>/[C] Dream 100 Client Database.md   (append only)
  3. lead-bank append    ->  <base>/[dream100-lead-bank].csv

Usage:
  python emit_outputs.py --leads leads.json --base "<root>" [--date YYYY-MM-DD]

leads.json: list of dicts with keys
  Name, Industry, Platforms, Instagram, YouTube, Email, ImprovementFound,
  OutreachAngle, ContactMethod, Fit ("green" | "yellow" | "red")
"""
import argparse, csv, datetime, json, os, re

COLS = ["#", "Name", "Industry", "Platforms", "Instagram", "YouTube", "Email(s)",
        "Improvement Found", "Outreach Angle (Loom Topic)", "Contact Method", "Fit"]
WIDTHS = [4, 20, 16, 13, 22, 22, 28, 34, 40, 15, 8]
FILL = {"green": "C6EFCE", "yellow": "FFEB9C", "red": "FFC7CE"}
DB_HEADER = ("| # | Name | Industry | Platform(s) | Instagram | YouTube | Email(s) | "
             "Improvement Found | Outreach Angle | Contact Method | Fit |\n"
             "|---|------|----------|-------------|-----------|---------|----------|"
             "-------------------|----------------|----------------|-----|\n")


def norm_handle(v):
    return (v or "").split(" ")[0].lstrip("@").strip().lower().replace("n/a", "")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--leads", required=True)
    ap.add_argument("--base", required=True)
    ap.add_argument("--date", default=datetime.date.today().isoformat())
    a = ap.parse_args()

    leads = json.load(open(a.leads, encoding="utf-8-sig"))
    db_path = os.path.join(a.base, "[C] Dream 100 Client Database.md")
    bank_path = os.path.join(a.base, "[dream100-lead-bank].csv")

    # Safety net: drop anything already in the bank (should never trigger)
    known = set()
    if os.path.exists(bank_path):
        for r in csv.DictReader(open(bank_path, encoding="utf-8-sig")):
            known |= {(r.get("Name") or "").strip().lower(),
                      (r.get("YouTube") or "").strip().lower(),
                      (r.get("Instagram") or "").strip().lower(),
                      (r.get("Email") or "").strip().lower()}
        known.discard("")
    fresh = [l for l in leads
             if l["Name"].strip().lower() not in known
             and (norm_handle(l.get("YouTube")) or "\0") not in known
             and (norm_handle(l.get("Instagram")) or "\0") not in known
             and ((l.get("Email") or "").strip().lower() or "\0") not in known]
    skipped = len(leads) - len(fresh)
    if not fresh:
        raise SystemExit(f"nothing to write — all {skipped} leads already in the bank")

    # ---- 1. xlsx
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Dream 100"
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
        c = ws.cell(1, j, h)
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A3A5C")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[c.column_letter].width = w
    ws.freeze_panes = "B2"
    for i, l in enumerate(fresh):
        row = i + 2
        ws.row_dimensions[row].height = 70
        fill = FILL.get((l.get("Fit") or "").lower()) or ("FFFFFF" if i % 2 == 0 else "EBF3FB")
        vals = [i + 1, l["Name"], l.get("Industry", ""), l.get("Platforms", ""),
                l.get("Instagram", ""), l.get("YouTube", ""), l.get("Email") or "N/A",
                l.get("ImprovementFound", ""), l.get("OutreachAngle", ""),
                l.get("ContactMethod", ""), l.get("Fit", "")]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row, j, v)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", fgColor=fill)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
    xlsx_path = os.path.join(a.base, f"[C] Dream 100 Leads - {a.date}.xlsx")
    wb.save(xlsx_path)

    # ---- 2. client DB append
    esc = lambda s: str(s).replace("|", "\\|").replace("\n", " ")
    if not os.path.exists(db_path):
        open(db_path, "w", encoding="utf-8").write("# Dream 100 Client Database\n\n" + DB_HEADER)
    lines = open(db_path, encoding="utf-8").readlines()
    nums = [int(m.group(1)) for l in lines if (m := re.match(r"\|\s*(\d+)\s*\|", l))]
    nxt = (max(nums) + 1) if nums else 1
    with open(db_path, "a", encoding="utf-8") as f:
        for i, l in enumerate(fresh):
            f.write(f"| {nxt + i} | {esc(l['Name'])} | {esc(l.get('Industry', ''))} | "
                    f"{esc(l.get('Platforms', ''))} | {esc(l.get('Instagram', ''))} | "
                    f"{esc(l.get('YouTube', ''))} | {esc(l.get('Email') or 'N/A')} | "
                    f"{esc(l.get('ImprovementFound', ''))} | {esc(l.get('OutreachAngle', ''))} | "
                    f"{esc(l.get('ContactMethod', ''))} | {esc(l.get('Fit', ''))} |\n")

    # ---- 3. lead bank append
    new_file = not os.path.exists(bank_path)
    with open(bank_path, "a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["Name", "Instagram", "YouTube", "Email"])
        if new_file:
            w.writeheader()
        for l in fresh:
            w.writerow({"Name": l["Name"].strip().lower(),
                        "Instagram": norm_handle(l.get("Instagram")),
                        "YouTube": norm_handle(l.get("YouTube")),
                        "Email": (l.get("Email") or "").strip().lower().replace("n/a", "")})
    total = sum(1 for _ in open(bank_path, encoding="utf-8-sig")) - 1

    print(f"xlsx: {xlsx_path}")
    print(f"appended {len(fresh)} leads (DB rows {nxt}-{nxt + len(fresh) - 1}); "
          f"skipped {skipped} already-known; bank total {total}")


if __name__ == "__main__":
    main()
